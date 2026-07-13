"""
코더 간 불일치 항목의 합의 코딩 파일을 생성
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import HUMAN_DIR
from labels import (
    SENTIMENT_VALUES,
    TARGET_VALUES,
    STANCE_VALUES,
    SARCASM_VALUES,
)


# =====================================================================
# 1. 파일 경로 설정
# =====================================================================
INPUT_PATH = HUMAN_DIR / "coding_disagreements.xlsx"
OUTPUT_PATH = HUMAN_DIR / "consensus_coding.xlsx"


# =====================================================================
# 2. 코딩 항목
# =====================================================================
TASK_COLUMNS = [
    "sentiment",
    "target",
    "stance",
    "is_sarcasm_mockery",
]

LABEL_VALUES = {
    "consensus_sentiment": SENTIMENT_VALUES,
    "consensus_target": TARGET_VALUES,
    "consensus_stance": STANCE_VALUES,
    "consensus_is_sarcasm_mockery": SARCASM_VALUES,
}


# =====================================================================
# 3. 불일치 자료 불러오기
# =====================================================================
def load_disagreements() -> pd.DataFrame:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(
            f"불일치 파일이 없습니다: {INPUT_PATH}"
        )

    df = pd.read_excel(
        INPUT_PATH,
        dtype=str,
    ).fillna("")

    required_columns = {
        "sample_id",
        "analysis_unit_id",
        "case_id",
        "case_name",
        "comment_type",
        "parent_text",
        "analysis_text",
        "coder1_sentiment",
        "coder2_sentiment",
        "coder1_target",
        "coder2_target",
        "coder1_stance",
        "coder2_stance",
        "coder1_is_sarcasm_mockery",
        "coder2_is_sarcasm_mockery",
        "coder1_note",
        "coder2_note",
    }

    missing_columns = required_columns - set(df.columns)

    if missing_columns:
        raise ValueError(
            f"필요한 열이 없습니다: {sorted(missing_columns)}"
        )

    return df


# =====================================================================
# 4. 합의 라벨 초기값 생성
# =====================================================================
def add_consensus_columns(
    df: pd.DataFrame,
) -> pd.DataFrame:
    df = df.copy()

    for task in TASK_COLUMNS:
        coder1_column = f"coder1_{task}"
        coder2_column = f"coder2_{task}"
        consensus_column = f"consensus_{task}"

        df[consensus_column] = ""

        match_mask = (
            df[coder1_column]
            == df[coder2_column]
        )

        df.loc[
            match_mask,
            consensus_column,
        ] = df.loc[
            match_mask,
            coder1_column,
        ]

    df["consensus_note"] = ""

    return df


# =====================================================================
# 5. 드롭다운 추가
# =====================================================================
def add_dropdown(
    sheet,
    column_name: str,
    values: list[str],
) -> None:
    header_index = {
        cell.value: cell.column
        for cell in sheet[1]
    }

    column_letter = get_column_letter(
        header_index[column_name]
    )

    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=True,
    )

    validation.error = (
        "드롭다운 목록의 값만 입력하십시오."
    )
    validation.errorTitle = "잘못된 라벨"

    sheet.add_data_validation(validation)

    validation.add(
        f"{column_letter}2:{column_letter}{sheet.max_row}"
    )


# =====================================================================
# 6. Excel 서식
# =====================================================================
def apply_style(sheet) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="BDD7EE",
    )

    input_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    header_index = {
        cell.value: cell.column
        for cell in sheet[1]
    }

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in sheet.iter_rows(
        min_row=2,
        max_row=sheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    consensus_columns = [
        "consensus_sentiment",
        "consensus_target",
        "consensus_stance",
        "consensus_is_sarcasm_mockery",
        "consensus_note",
    ]

    for column_name in consensus_columns:
        column_number = header_index[column_name]

        for row_number in range(
            2,
            sheet.max_row + 1,
        ):
            sheet.cell(
                row=row_number,
                column=column_number,
            ).fill = input_fill

    width_map = {
        "sample_id": 12,
        "analysis_unit_id": 18,
        "case_id": 10,
        "case_name": 30,
        "comment_type": 15,
        "parent_text": 55,
        "analysis_text": 55,
        "coder1_sentiment": 18,
        "coder2_sentiment": 18,
        "consensus_sentiment": 22,
        "coder1_target": 18,
        "coder2_target": 18,
        "consensus_target": 22,
        "coder1_stance": 18,
        "coder2_stance": 18,
        "consensus_stance": 22,
        "coder1_is_sarcasm_mockery": 25,
        "coder2_is_sarcasm_mockery": 25,
        "consensus_is_sarcasm_mockery": 28,
        "coder1_note": 35,
        "coder2_note": 35,
        "consensus_note": 40,
    }

    for column_name, width in width_map.items():
        column_letter = get_column_letter(
            header_index[column_name]
        )

        sheet.column_dimensions[
            column_letter
        ].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


# =====================================================================
# 7. 합의 코딩 파일 생성
# =====================================================================
def create_workbook(
    df: pd.DataFrame,
) -> None:
    output_columns = [
        "sample_id",
        "analysis_unit_id",
        "case_id",
        "case_name",
        "comment_type",
        "parent_text",
        "analysis_text",
        "coder1_sentiment",
        "coder2_sentiment",
        "consensus_sentiment",
        "coder1_target",
        "coder2_target",
        "consensus_target",
        "coder1_stance",
        "coder2_stance",
        "consensus_stance",
        "coder1_is_sarcasm_mockery",
        "coder2_is_sarcasm_mockery",
        "consensus_is_sarcasm_mockery",
        "coder1_note",
        "coder2_note",
        "consensus_note",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consensus"

    sheet.append(output_columns)

    for _, row in df.iterrows():
        sheet.append(
            [
                row[column]
                for column in output_columns
            ]
        )

    apply_style(sheet)

    for column_name, values in LABEL_VALUES.items():
        add_dropdown(
            sheet,
            column_name,
            values,
        )

    workbook.save(OUTPUT_PATH)


# =====================================================================
# 8. 실행
# =====================================================================
def main() -> None:
    if OUTPUT_PATH.exists():
        raise FileExistsError(
            "consensus_coding.xlsx가 이미 존재합니다."
        )

    df = load_disagreements()
    df = add_consensus_columns(df)

    create_workbook(df)

    print("=" * 60)
    print("합의 코딩 Excel 생성 완료")
    print("=" * 60)
    print(f"불일치 표본 수: {len(df)}개")
    print(f"파일: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()