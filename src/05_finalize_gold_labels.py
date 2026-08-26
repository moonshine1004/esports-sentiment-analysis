"""합의 결과를 반영하되 나머지 코더1 라벨을 보존해 300건 골드라벨 생성."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import HUMAN_DIR, HUMAN_SAMPLE_MAX, INTERCODER_SAMPLE_SIZE
from intercoder import (
    METADATA_COLUMNS,
    TASK_COLUMNS,
    add_item_agreement_columns,
    align_overlap,
    load_coder,
    load_intercoder_manifest,
    verify_coder1_hash,
)
from labels import TASK_LABELS
from text_utils import normalize_boolean_text, normalize_label

CODER_PATHS = {f"coder{i}": HUMAN_DIR / f"coder{i}_coding.xlsx" for i in (1, 2, 3)}
MANIFEST_PATH = HUMAN_DIR / "intercoder_sample_manifest.json"
ADJUDICATION_PATH = HUMAN_DIR / "adjudication_workbook.xlsx"
OUTPUT_CSV_PATH = HUMAN_DIR / "gold_labels.csv"
OUTPUT_XLSX_PATH = HUMAN_DIR / "gold_labels.xlsx"
CODER_NAMES = ["coder1", "coder2", "coder3"]


def normalize_and_validate_decisions(
    decisions: pd.DataFrame,
    items: pd.DataFrame,
) -> pd.DataFrame:
    final_columns = [f"final_{task}" for task in TASK_COLUMNS]
    required_columns = {"sample_id"} | set(final_columns)
    missing = required_columns - set(decisions.columns)
    if missing:
        raise ValueError(f"합의 파일에 필요한 열이 없습니다: {sorted(missing)}")
    if decisions["sample_id"].duplicated().any():
        duplicate_ids = decisions.loc[
            decisions["sample_id"].duplicated(keep=False), "sample_id"
        ].tolist()[:10]
        raise ValueError(f"합의 파일에 중복 sample_id가 있습니다: {duplicate_ids}")

    decisions = decisions.copy()
    decisions["sample_id"] = decisions["sample_id"].apply(lambda value: str(value).strip())
    expected_ids = set(items.loc[items["requires_adjudication"], "sample_id"])
    actual_ids = set(decisions["sample_id"])
    if actual_ids != expected_ids:
        missing_ids = sorted(expected_ids - actual_ids)[:10]
        extra_ids = sorted(actual_ids - expected_ids)[:10]
        raise ValueError(
            f"합의 파일의 항목이 실제 불일치 항목과 다릅니다: 누락={missing_ids}, 추가={extra_ids}"
        )

    item_lookup = items.set_index("sample_id")
    for task in TASK_COLUMNS:
        column = f"final_{task}"
        normalizer = normalize_boolean_text if task == "is_sarcasm_mockery" else normalize_label
        decisions[column] = decisions[column].apply(normalizer)
        invalid = ~decisions[column].isin(TASK_LABELS[task])
        if invalid.any():
            bad = decisions.loc[invalid, "sample_id"].tolist()[:10]
            raise ValueError(f"{column}에 세 코더가 합의한 최종값을 입력해야 합니다: {bad}")

        for _, row in decisions.iterrows():
            sample_id = row["sample_id"]
            if bool(item_lookup.loc[sample_id, f"agreement_{task}"]):
                unanimous_value = item_lookup.loc[sample_id, f"coder1_{task}"]
                if row[column] != unanimous_value:
                    raise ValueError(
                        f"이미 일치한 {task} 값은 변경할 수 없습니다: "
                        f"{sample_id}, expected={unanimous_value}, actual={row[column]}"
                    )
    return decisions


def build_gold_labels(
    coder1: pd.DataFrame,
    items: pd.DataFrame,
    decisions: pd.DataFrame,
) -> pd.DataFrame:
    result = coder1.copy().set_index("sample_id")
    item_lookup = items.set_index("sample_id")
    decision_lookup = decisions.set_index("sample_id")
    overlap_ids = item_lookup.index.tolist()

    for task in TASK_COLUMNS:
        source_column = f"gold_{task}_source"
        result[source_column] = "coder1_only"
        agreement_mask = item_lookup[f"agreement_{task}"].astype(bool)
        unanimous_ids = item_lookup.index[agreement_mask].tolist()
        disagreement_ids = item_lookup.index[~agreement_mask].tolist()
        result.loc[unanimous_ids, source_column] = "unanimous_3_coders"
        if disagreement_ids:
            final_values = decision_lookup.loc[disagreement_ids, f"final_{task}"]
            result.loc[disagreement_ids, task] = final_values.to_numpy()
            result.loc[disagreement_ids, source_column] = "adjudicated_3_coders"

    result["in_intercoder_sample"] = result.index.isin(overlap_ids)
    result = result.reset_index()
    output_columns = ["sample_id"] + METADATA_COLUMNS + TASK_COLUMNS
    source_columns = [f"gold_{task}_source" for task in TASK_COLUMNS]
    output = result[output_columns + ["in_intercoder_sample"] + source_columns].copy()
    output = output.rename(columns={task: f"gold_{task}" for task in TASK_COLUMNS})
    return output


def style_gold_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    header_fill = PatternFill("solid", fgColor="4472C4")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "sample_id": 12,
        "analysis_unit_id": 18,
        "case_id": 9,
        "case_name": 30,
        "comment_type": 14,
        "parent_text": 50,
        "analysis_text": 60,
        "in_intercoder_sample": 22,
    }
    headers = {cell.value: cell.column for cell in sheet[1]}
    for name, column in headers.items():
        width = widths.get(name, 24 if str(name).endswith("_source") else 20)
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def finalize_gold_labels() -> pd.DataFrame:
    if not ADJUDICATION_PATH.exists():
        raise FileNotFoundError(
            "먼저 05_analyze_intercoder_reliability.py를 실행하고 합의 파일을 완성하십시오."
        )
    manifest = load_intercoder_manifest(MANIFEST_PATH)
    if manifest["sample_size"] != INTERCODER_SAMPLE_SIZE:
        raise ValueError("공통 표본 매니페스트와 INTERCODER_SAMPLE_SIZE가 다릅니다.")
    verify_coder1_hash(CODER_PATHS["coder1"], manifest["coder1_sha256"])
    coders = {
        "coder1": load_coder(CODER_PATHS["coder1"], "coder1", expected_rows=HUMAN_SAMPLE_MAX),
        "coder2": load_coder(CODER_PATHS["coder2"], "coder2", expected_rows=INTERCODER_SAMPLE_SIZE),
        "coder3": load_coder(CODER_PATHS["coder3"], "coder3", expected_rows=INTERCODER_SAMPLE_SIZE),
    }
    overlap = align_overlap(coders, manifest["sample_ids"])
    items = add_item_agreement_columns(overlap, CODER_NAMES)
    decisions = pd.read_excel(
        ADJUDICATION_PATH,
        sheet_name="Adjudication",
        dtype=str,
    ).fillna("")
    decisions = normalize_and_validate_decisions(decisions, items)
    output = build_gold_labels(coders["coder1"], items, decisions)
    output.to_csv(OUTPUT_CSV_PATH, index=False, encoding="utf-8-sig")
    output.to_excel(OUTPUT_XLSX_PATH, index=False, sheet_name="Gold labels")
    style_gold_workbook(OUTPUT_XLSX_PATH)
    return output


def main() -> None:
    output = finalize_gold_labels()
    adjudicated_counts = {
        task: int((output[f"gold_{task}_source"] == "adjudicated_3_coders").sum())
        for task in TASK_COLUMNS
    }
    print(f"합의 반영 인간 기준 라벨 {len(output)}건 생성: {OUTPUT_XLSX_PATH}")
    print(f"과업별 합의 조정 건수: {adjudicated_counts}")


if __name__ == "__main__":
    main()
