"""세 코더의 동일 100건에 대해 합의 전 신뢰도와 불일치 합의 파일 생성."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import HUMAN_DIR, HUMAN_SAMPLE_MAX, INTERCODER_SAMPLE_SIZE
from intercoder import (
    METADATA_COLUMNS,
    TASK_COLUMNS,
    add_item_agreement_columns,
    align_overlap,
    calculate_reliability,
    load_coder,
    load_intercoder_manifest,
    verify_coder1_hash,
)
from labels import TASK_LABELS

CODER_PATHS = {f"coder{i}": HUMAN_DIR / f"coder{i}_coding.xlsx" for i in (1, 2, 3)}
MANIFEST_PATH = HUMAN_DIR / "intercoder_sample_manifest.json"
SUMMARY_CSV = HUMAN_DIR / "intercoder_reliability_summary.csv"
PAIRWISE_CSV = HUMAN_DIR / "intercoder_reliability_pairwise.csv"
REPORT_XLSX = HUMAN_DIR / "intercoder_reliability.xlsx"
ITEMS_CSV = HUMAN_DIR / "intercoder_agreement_items.csv"
ADJUDICATION_XLSX = HUMAN_DIR / "adjudication_workbook.xlsx"

CODER_NAMES = ["coder1", "coder2", "coder3"]


def build_adjudication_rows(items: pd.DataFrame) -> pd.DataFrame:
    output = items.copy()
    for task in TASK_COLUMNS:
        first_coder_column = f"coder1_{task}"
        output[f"final_{task}"] = output[first_coder_column].where(
            output[f"agreement_{task}"],
            "",
        )
    output["adjudication_note"] = ""
    output["adjudicated_by"] = ""
    output["adjudicated_at"] = ""
    output = output.loc[output["requires_adjudication"]].copy()

    ordered_columns = ["sample_id"] + METADATA_COLUMNS + ["disagreement_tasks"]
    for task in TASK_COLUMNS:
        ordered_columns.extend(
            [
                f"coder1_{task}",
                f"coder2_{task}",
                f"coder3_{task}",
                f"agreement_{task}",
                f"final_{task}",
            ]
        )
    ordered_columns.extend(["adjudication_note", "adjudicated_by", "adjudicated_at"])
    return output[ordered_columns]


def style_adjudication_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook["Adjudication"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    final_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for task in TASK_COLUMNS:
        column = headers[f"final_{task}"]
        letter = get_column_letter(column)
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(TASK_LABELS[task])}"',
            allow_blank=False,
        )
        validation.error = "세 코더가 합의한 허용 라벨을 선택하십시오."
        validation.errorTitle = "잘못된 합의 라벨"
        validation.showErrorMessage = True
        sheet.add_data_validation(validation)
        if sheet.max_row >= 2:
            validation.add(f"{letter}2:{letter}{sheet.max_row}")
            for row_number in range(2, sheet.max_row + 1):
                sheet.cell(row_number, column).fill = final_fill

    width_by_name = {
        "sample_id": 12,
        "analysis_unit_id": 18,
        "case_id": 9,
        "case_name": 28,
        "comment_type": 14,
        "parent_text": 50,
        "analysis_text": 60,
        "disagreement_tasks": 30,
        "adjudication_note": 45,
        "adjudicated_by": 18,
        "adjudicated_at": 20,
    }
    for name, column in headers.items():
        width = width_by_name.get(name, 20 if str(name).startswith("final_") else 18)
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["항목", "내용"])
    instructions.append(["목적", "합의 전 신뢰도 계산에서 불일치한 항목만 조정합니다."])
    instructions.append(["final_*", "노란색 셀에 세 코더가 논의해 합의한 최종 라벨을 입력합니다."])
    instructions.append(["이미 일치한 과업", "해당 final_* 값은 자동 입력되며 변경하지 않습니다."])
    instructions.append(["기록", "합의 근거, 참여자, 합의 일시를 오른쪽 세 열에 기록합니다."])
    for cell in instructions[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 90
    instructions.sheet_view.showGridLines = False
    workbook.active = workbook.sheetnames.index("Adjudication")
    workbook.save(path)


def create_adjudication_workbook(items: pd.DataFrame, path: Path) -> pd.DataFrame:
    adjudication = build_adjudication_rows(items)
    adjudication.to_excel(path, index=False, sheet_name="Adjudication")
    style_adjudication_workbook(path)
    return adjudication


def style_reliability_report(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="4472C4")
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = min(
                max(len(str(cell.value or "")) for cell in column_cells) + 2,
                55,
            )
            sheet.column_dimensions[letter].width = max(12, max_length)

    summary = workbook["Summary"]
    summary_headers = {cell.value: cell.column for cell in summary[1]}
    for name in ["unanimous_agreement", "fleiss_kappa", "krippendorff_alpha_nominal"]:
        column = summary_headers[name]
        for row in range(2, summary.max_row + 1):
            summary.cell(row, column).number_format = "0.000"
    percent_column = summary_headers["percent_unanimous_agreement"]
    for row in range(2, summary.max_row + 1):
        summary.cell(row, percent_column).number_format = '0.0"%"'

    pairwise = workbook["Pairwise"]
    pair_headers = {cell.value: cell.column for cell in pairwise[1]}
    for name in ["agreement_rate", "cohen_kappa"]:
        column = pair_headers[name]
        for row in range(2, pairwise.max_row + 1):
            pairwise.cell(row, column).number_format = "0.000"
    percent_column = pair_headers["percent_agreement"]
    for row in range(2, pairwise.max_row + 1):
        pairwise.cell(row, percent_column).number_format = '0.0"%"'
    workbook.save(path)


def save_reliability_outputs(
    summary: pd.DataFrame,
    pairwise: pd.DataFrame,
    items: pd.DataFrame,
    manifest: dict,
) -> None:
    summary.to_csv(SUMMARY_CSV, index=False, encoding="utf-8-sig")
    pairwise.to_csv(PAIRWISE_CSV, index=False, encoding="utf-8-sig")
    items.to_csv(ITEMS_CSV, index=False, encoding="utf-8-sig")
    metadata = pd.DataFrame(
        [
            {"item": "calculation_stage", "value": "before adjudication"},
            {"item": "sample_size", "value": manifest["sample_size"]},
            {"item": "n_coders", "value": len(CODER_NAMES)},
            {"item": "coder1_sha256", "value": manifest["coder1_sha256"]},
            {"item": "random_seed", "value": manifest.get("random_seed", "")},
            {"item": "missing_values", "value": "not allowed"},
        ]
    )
    with pd.ExcelWriter(REPORT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        pairwise.to_excel(writer, sheet_name="Pairwise", index=False)
        items.to_excel(writer, sheet_name="Item agreement", index=False)
        metadata.to_excel(writer, sheet_name="Run metadata", index=False)
    style_reliability_report(REPORT_XLSX)


def analyze_intercoder_reliability(
    *,
    overwrite_adjudication: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if ADJUDICATION_XLSX.exists() and not overwrite_adjudication:
        raise FileExistsError(
            "기존 합의 입력을 보호하기 위해 adjudication_workbook.xlsx를 덮어쓰지 않습니다. "
            "아직 합의를 시작하지 않았고 다시 만들려는 경우에만 --overwrite-adjudication을 사용하십시오."
        )
    manifest = load_intercoder_manifest(MANIFEST_PATH)
    if manifest["sample_size"] != INTERCODER_SAMPLE_SIZE:
        raise ValueError(
            "매니페스트의 공통 표본 수와 INTERCODER_SAMPLE_SIZE가 다릅니다: "
            f"{manifest['sample_size']} != {INTERCODER_SAMPLE_SIZE}"
        )
    verify_coder1_hash(CODER_PATHS["coder1"], manifest["coder1_sha256"])
    coders = {
        "coder1": load_coder(CODER_PATHS["coder1"], "coder1", expected_rows=HUMAN_SAMPLE_MAX),
        "coder2": load_coder(CODER_PATHS["coder2"], "coder2", expected_rows=INTERCODER_SAMPLE_SIZE),
        "coder3": load_coder(CODER_PATHS["coder3"], "coder3", expected_rows=INTERCODER_SAMPLE_SIZE),
    }
    overlap = align_overlap(coders, manifest["sample_ids"])
    summary, pairwise = calculate_reliability(overlap, CODER_NAMES)
    items = add_item_agreement_columns(overlap, CODER_NAMES)
    save_reliability_outputs(summary, pairwise, items, manifest)
    adjudication = create_adjudication_workbook(items, ADJUDICATION_XLSX)
    return summary, pairwise, adjudication


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--overwrite-adjudication",
        action="store_true",
        help="기존 합의 워크북을 지웁니다. 합의를 시작하지 않은 경우에만 사용하십시오.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    summary, _, adjudication = analyze_intercoder_reliability(
        overwrite_adjudication=args.overwrite_adjudication
    )
    print(summary.to_string(index=False))
    print(f"\n신뢰도 보고서: {REPORT_XLSX}")
    print(f"합의가 필요한 항목: {len(adjudication)}건")
    print(f"합의 입력 파일: {ADJUDICATION_XLSX}")


if __name__ == "__main__":
    main()
