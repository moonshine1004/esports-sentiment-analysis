"""기존 코더 1을 보존하고 동일한 공통 100건으로 코더 2·3 파일 생성."""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import HUMAN_DIR, HUMAN_SAMPLE_MAX, INTERCODER_SAMPLE_SIZE, RANDOM_SEED
from intercoder import load_coder, sha256_file
from labels import SARCASM_VALUES, SENTIMENT_VALUES, STANCE_VALUES, TARGET_VALUES

CODER1_PATH = HUMAN_DIR / "coder1_coding.xlsx"
CODER2_PATH = HUMAN_DIR / "coder2_coding.xlsx"
CODER3_PATH = HUMAN_DIR / "coder3_coding.xlsx"
OVERLAP_PATH = HUMAN_DIR / "intercoder_sample.csv"
MANIFEST_PATH = HUMAN_DIR / "intercoder_sample_manifest.json"

GROUP_COLUMNS = ["case_id", "comment_type"]
OUTPUT_COLUMNS = [
    "sample_id",
    "analysis_unit_id",
    "case_id",
    "case_name",
    "comment_type",
    "parent_text",
    "analysis_text",
    "sentiment",
    "target",
    "stance",
    "is_sarcasm_mockery",
    "coder_note",
]


def allocate_stratified_counts(
    df: pd.DataFrame,
    sample_size: int,
    group_columns: list[str] | None = None,
) -> pd.DataFrame:
    """층별 최소 1건과 최대용량을 지키는 최대잔여법 배분."""
    group_columns = group_columns or GROUP_COLUMNS
    if not 1 <= sample_size <= len(df):
        raise ValueError(f"sample_size는 1 이상 {len(df)} 이하여야 합니다.")
    group_sizes = (
        df.groupby(group_columns, dropna=False, sort=True)
        .size()
        .reset_index(name="group_size")
    )
    if group_sizes.empty:
        raise ValueError("층화 표본을 추출할 자료가 없습니다.")

    group_sizes["sample_n"] = 0
    if sample_size >= len(group_sizes):
        group_sizes["sample_n"] = 1
    else:
        largest = group_sizes.sort_values(
            ["group_size"] + group_columns,
            ascending=[False] + [True] * len(group_columns),
        ).head(sample_size)
        group_sizes.loc[largest.index, "sample_n"] = 1

    remaining = sample_size - int(group_sizes["sample_n"].sum())
    if remaining:
        group_sizes["capacity"] = group_sizes["group_size"] - group_sizes["sample_n"]
        total_capacity = int(group_sizes["capacity"].sum())
        quota = group_sizes["capacity"] / total_capacity * remaining
        additional = quota.apply(math.floor).astype(int)
        group_sizes["sample_n"] += additional
        group_sizes["remainder"] = quota - additional
        leftover = sample_size - int(group_sizes["sample_n"].sum())
        available = group_sizes[
            group_sizes["sample_n"] < group_sizes["group_size"]
        ].sort_values(
            ["remainder", "group_size"] + group_columns,
            ascending=[False, False] + [True] * len(group_columns),
        )
        for index in available.index[:leftover]:
            group_sizes.loc[index, "sample_n"] += 1

    if int(group_sizes["sample_n"].sum()) != sample_size:
        raise RuntimeError("층별 공통 표본 수 배분에 실패했습니다.")
    if (group_sizes["sample_n"] > group_sizes["group_size"]).any():
        raise RuntimeError("층의 가용 표본 수보다 많은 수가 배분되었습니다.")
    return group_sizes[group_columns + ["group_size", "sample_n"]]


def select_overlap_sample(
    coder1: pd.DataFrame,
    sample_size: int,
    random_seed: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """코더1 300건에서 사례×댓글유형 비율을 반영한 공통 표본 추출."""
    counts = allocate_stratified_counts(coder1, sample_size)
    sampled_groups = []
    for offset, group in counts.iterrows():
        mask = pd.Series(True, index=coder1.index)
        for column in GROUP_COLUMNS:
            mask &= coder1[column].eq(group[column])
        group_df = coder1.loc[mask]
        sample_n = int(group["sample_n"])
        if sample_n:
            sampled_groups.append(
                group_df.sample(n=sample_n, random_state=random_seed + int(offset))
            )
    overlap = pd.concat(sampled_groups, ignore_index=True)
    overlap = overlap.sample(frac=1, random_state=random_seed).reset_index(drop=True)
    if len(overlap) != sample_size or overlap["sample_id"].duplicated().any():
        raise RuntimeError("공통 표본 추출 결과가 요청한 크기 또는 고유성 조건과 다릅니다.")
    overlap.insert(0, "selection_order", range(1, len(overlap) + 1))
    return overlap, counts


def add_dropdown(sheet, column_name: str, values: list[str]) -> None:
    header_index = {cell.value: cell.column for cell in sheet[1]}
    column_letter = get_column_letter(header_index[column_name])
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=True,
    )
    validation.error = "드롭다운 목록에 있는 값만 입력하십시오."
    validation.errorTitle = "잘못된 라벨"
    validation.showErrorMessage = True
    sheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{sheet.max_row}")


def apply_style(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
    input_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    input_columns = {
        "sentiment",
        "target",
        "stance",
        "is_sarcasm_mockery",
        "coder_note",
    }
    header_index = {cell.value: cell.column for cell in sheet[1]}
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_name in input_columns:
        column_number = header_index[column_name]
        for row_number in range(2, sheet.max_row + 1):
            sheet.cell(row=row_number, column=column_number).fill = input_fill
    width_map = {
        "sample_id": 12,
        "analysis_unit_id": 18,
        "case_id": 10,
        "case_name": 32,
        "comment_type": 15,
        "parent_text": 60,
        "analysis_text": 60,
        "sentiment": 16,
        "target": 18,
        "stance": 18,
        "is_sarcasm_mockery": 22,
        "coder_note": 40,
    }
    for column_name, width in width_map.items():
        letter = get_column_letter(header_index[column_name])
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def add_instructions_sheet(workbook: Workbook, coder_name: str) -> None:
    sheet = workbook.create_sheet("Instructions")
    rows = [
        ("코더", coder_name),
        ("대상", "코더 1의 300건 중 세 코더 공통 100건"),
        ("독립 코딩", "신뢰도 산출 전에는 다른 코더의 결과를 보거나 협의하지 않습니다."),
        ("입력 열", "노란색 sentiment, target, stance, is_sarcasm_mockery, coder_note"),
        ("sentiment", ", ".join(SENTIMENT_VALUES)),
        ("target", ", ".join(TARGET_VALUES)),
        ("stance", ", ".join(STANCE_VALUES)),
        ("is_sarcasm_mockery", ", ".join(SARCASM_VALUES)),
        ("세부 기준", "README.md와 docs/coding_protocol.md의 고정된 코딩 규칙을 따릅니다."),
    ]
    for row in rows:
        sheet.append(row)
    for row in sheet.iter_rows():
        row[0].font = Font(bold=True)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 90
    sheet.sheet_view.showGridLines = False


def create_workbook(df: pd.DataFrame, output_path: Path, coder_name: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Coding"
    sheet.append(OUTPUT_COLUMNS)
    for _, row in df.iterrows():
        sheet.append(
            [
                row["sample_id"],
                row["analysis_unit_id"],
                row["case_id"],
                row["case_name"],
                row["comment_type"],
                row["parent_text"],
                row["analysis_text"],
                "",
                "",
                "",
                "",
                "",
            ]
        )
    apply_style(sheet)
    add_dropdown(sheet, "sentiment", SENTIMENT_VALUES)
    add_dropdown(sheet, "target", TARGET_VALUES)
    add_dropdown(sheet, "stance", STANCE_VALUES)
    add_dropdown(sheet, "is_sarcasm_mockery", SARCASM_VALUES)
    add_instructions_sheet(workbook, coder_name)
    workbook.active = workbook.sheetnames.index("Coding")
    workbook.save(output_path)


def write_manifest(
    overlap: pd.DataFrame,
    counts: pd.DataFrame,
    coder1_sha256: str,
) -> None:
    manifest = {
        "schema_version": 1,
        "coder1_file": CODER1_PATH.name,
        "coder1_sha256": coder1_sha256,
        "human_sample_size": HUMAN_SAMPLE_MAX,
        "sample_size": len(overlap),
        "random_seed": RANDOM_SEED,
        "stratification_columns": GROUP_COLUMNS,
        "sample_ids": overlap["sample_id"].tolist(),
        "strata": counts.to_dict(orient="records"),
    }
    MANIFEST_PATH.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def generate_coder_workbooks(*, overwrite_generated: bool = False) -> pd.DataFrame:
    if not CODER1_PATH.exists():
        raise FileNotFoundError(
            "기존 coder1_coding.xlsx가 필요합니다. 이 스크립트는 해당 파일을 생성하거나 덮어쓰지 않습니다."
        )
    generated_paths = [CODER2_PATH, CODER3_PATH, OVERLAP_PATH, MANIFEST_PATH]
    existing = [path for path in generated_paths if path.exists()]
    if existing and not overwrite_generated:
        raise FileExistsError(
            "기존 생성 파일을 보호하기 위해 중단합니다. 다시 만들려면 "
            f"--overwrite-generated를 사용하십시오: {[str(path) for path in existing]}"
        )

    coder1_hash_before = sha256_file(CODER1_PATH)
    coder1 = load_coder(CODER1_PATH, "coder1", expected_rows=HUMAN_SAMPLE_MAX)
    overlap, counts = select_overlap_sample(
        coder1,
        sample_size=INTERCODER_SAMPLE_SIZE,
        random_seed=RANDOM_SEED,
    )
    create_workbook(overlap, CODER2_PATH, "coder2")
    create_workbook(overlap, CODER3_PATH, "coder3")
    overlap[
        ["selection_order", "sample_id", "analysis_unit_id", "case_id", "case_name", "comment_type"]
    ].to_csv(OVERLAP_PATH, index=False, encoding="utf-8-sig")
    write_manifest(overlap, counts, coder1_hash_before)

    coder1_hash_after = sha256_file(CODER1_PATH)
    if coder1_hash_after != coder1_hash_before:
        raise RuntimeError("coder1_coding.xlsx의 해시가 실행 중 변경되었습니다.")
    return overlap


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--overwrite-generated",
        action="store_true",
        help="생성물인 coder2/3·공통표본·매니페스트만 다시 만듭니다. coder1은 항상 보존합니다.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    overlap = generate_coder_workbooks(overwrite_generated=args.overwrite_generated)
    print("=" * 60)
    print("코더 2·3 공통 표본 Excel 생성 완료 (코더 1 파일은 변경하지 않음)")
    print("=" * 60)
    print(f"공통 코딩 표본: {len(overlap)}개")
    print(f"코더 파일: {CODER2_PATH}, {CODER3_PATH}")
    print(f"매니페스트: {MANIFEST_PATH}")


if __name__ == "__main__":
    main()
